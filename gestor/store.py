"""Base de datos (JSON) + generador del Excel para Power Automate."""
import calendar
import json
import os
from datetime import date, timedelta

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

from . import config

# Columnas del Excel (nombre visible) -> clave interna ('_' = calculada)
COLS = [
    ("ID", "hash"),
    ("Nombre", "nombre"),
    ("NumeroFiscal", "numero_fiscal"),
    ("TipoDocumento", "tipo_documento"),
    ("Pais", "pais"),
    ("FechaSello", "fecha_sello"),
    ("FechaInicio", "fecha_inicio"),
    ("FechaFinOficial", "fecha_fin"),
    ("CaducidadEstimada", "_estimada"),
    ("CaducidadEfectiva", "_efectiva"),
    ("FechaAviso", "_fecha_aviso"),
    ("DiasParaCaducar", "_dias"),
    ("EstadoCaducidad", "_estado_cad"),
    ("EnAviso", "_en_aviso"),
    ("Estado", "estado"),
    ("Motor", "motor"),
    ("Observaciones", "observaciones"),
    ("Archivos", "_archivos"),
    ("FechaAnalisis", "fecha_analisis"),
]


def load_db():
    if os.path.exists(config.DB_PATH):
        try:
            with open(config.DB_PATH, "r", encoding="utf-8") as f:
                return json.load(f)
        except Exception:
            return []
    return []


def save_db(records):
    os.makedirs(config.APP_DIR, exist_ok=True)
    with open(config.DB_PATH, "w", encoding="utf-8") as f:
        json.dump(records, f, ensure_ascii=False, indent=2)


def has_hash(records, h):
    return any(r.get("hash") == h for r in records)


def upsert(records, record):
    for i, r in enumerate(records):
        if r.get("hash") == record.get("hash"):
            records[i] = record
            return records
    records.append(record)
    return records


def add_months(iso, months):
    if not iso:
        return ""
    try:
        y, m, d = map(int, iso.split("-"))
    except Exception:
        return ""
    total = m - 1 + int(months)
    y2 = y + total // 12
    m2 = total % 12 + 1
    d2 = min(d, calendar.monthrange(y2, m2)[1])
    return date(y2, m2, d2).isoformat()


def add_days(iso, days):
    if not iso:
        return ""
    try:
        return (date.fromisoformat(iso) + timedelta(days=days)).isoformat()
    except Exception:
        return ""


def _is_residencia(rec):
    return "residencia" in (rec.get("tipo_documento", "") or "").lower()


def caducidad_info(rec, cfg):
    """Calcula caducidad y aviso segun los ajustes del usuario."""
    meses = cfg.get("caducidad_meses", 12)
    dias_aviso = cfg.get("dias_aviso", 30)
    base_mode = cfg.get("base_caducidad", "efectiva")
    solo_res = cfg.get("aviso_solo_residencia", False)

    estimada = add_months(rec.get("fecha_sello", ""), meses)
    fin = rec.get("fecha_fin", "") or ""
    if base_mode == "fin":
        base = fin
    elif base_mode == "estimada":
        base = estimada
    else:
        base = fin or estimada
    if solo_res and not _is_residencia(rec):
        base = ""

    info = {"estimada": estimada, "efectiva": base, "fecha_aviso": "",
            "dias": "", "estado_cad": "", "en_aviso": ""}
    if base:
        info["fecha_aviso"] = add_days(base, -dias_aviso)
        try:
            d = (date.fromisoformat(base) - date.today()).days
            info["dias"] = d
            if d < 0:
                info["estado_cad"] = "Caducado"
            elif d <= dias_aviso:
                info["estado_cad"] = "Por caducar"
            else:
                info["estado_cad"] = "Vigente"
            info["en_aviso"] = "SI" if d <= dias_aviso else "NO"
        except Exception:
            pass
    return info


def _row_values(rec, cfg):
    ci = caducidad_info(rec, cfg)
    calc = {
        "_estimada": ci["estimada"],
        "_efectiva": ci["efectiva"],
        "_fecha_aviso": ci["fecha_aviso"],
        "_dias": ci["dias"],
        "_estado_cad": ci["estado_cad"],
        "_en_aviso": ci["en_aviso"],
        "_archivos": " | ".join(rec.get("archivos", []) or []),
    }
    out = []
    for _, key in COLS:
        out.append(calc[key] if key in calc else rec.get(key, ""))
    return out


def _sort_key(rec, cfg):
    return caducidad_info(rec, cfg)["efectiva"] or "9999-12-31"


def write_excel(records, path, cfg):
    wb = Workbook()
    ws = wb.active
    ws.title = "Certificados"
    FONT = "Calibri"
    navy = "1F4E78"
    headers = [h for h, _ in COLS]

    for c, h in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=c, value=h)
        cell.font = Font(name=FONT, size=10, bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor=navy)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    ordered = sorted(records, key=lambda r: _sort_key(r, cfg))
    thin = Side(style="thin", color="D9D9D9")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    wrap_cols = {"TipoDocumento", "Observaciones", "Archivos", "Nombre"}
    for i, rec in enumerate(ordered):
        for c, v in enumerate(_row_values(rec, cfg), start=1):
            cell = ws.cell(row=i + 2, column=c, value=v)
            cell.font = Font(name=FONT, size=10)
            cell.alignment = Alignment(vertical="center", wrap_text=headers[c - 1] in wrap_cols)
            cell.border = border

    widths = [14, 24, 22, 32, 14, 11, 11, 13, 15, 15, 12, 11, 14, 9, 12, 14, 36, 30, 16]
    for idx, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = w

    nrows = max(len(ordered) + 1, 2)
    ref = f"A1:{get_column_letter(len(headers))}{nrows}"
    table = Table(displayName="Certificados", ref=ref)
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2", showRowStripes=True, showColumnStripes=False)
    ws.add_table(table)
    ws.freeze_panes = "A2"
    ws.sheet_view.showGridLines = False

    os.makedirs(os.path.dirname(os.path.abspath(path)), exist_ok=True)
    wb.save(path)
